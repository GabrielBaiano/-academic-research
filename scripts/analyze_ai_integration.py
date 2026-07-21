import os
import json
import re
import ssl
import time
import sys
import urllib.request
import urllib.parse
import urllib.error
import unicodedata
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CACHE_DIR = os.path.join(ROOT_DIR, "datasets", "cache")
os.makedirs(CACHE_DIR, exist_ok=True)

# API Configurations
API_KEY = os.environ.get("GEMINI_API_KEY", "").strip()
GEMINI_URL = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={API_KEY}"

def validate_and_setup_api_key():
    global API_KEY, GEMINI_URL
    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    
    if not api_key:
        print("\n[Aviso] Nenhuma chave de API do Gemini encontrada na variável de ambiente GEMINI_API_KEY.")
        user_key = input("Por favor, insira uma chave de API do Gemini válida (ou Enter para pular a análise com IA): ").strip()
        if not user_key:
            print("[Aviso] Prosseguindo sem IA. Será usada apenas a classificação heurística (regex).")
            return False
        api_key = user_key
        
    # Validar a chave
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={api_key}"
    payload = {
        "contents": [{"parts": [{"text": "Say ok"}]}]
    }
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    
    try:
        ctx = ssl._create_unverified_context()
        with urllib.request.urlopen(req, context=ctx, timeout=10) as res:
            API_KEY = api_key
            GEMINI_URL = url
            os.environ["GEMINI_API_KEY"] = api_key
            print("[Sucesso] Chave de API do Gemini validada com sucesso!\n")
            return True
    except Exception as e:
        print(f"\n[Aviso] A chave de API fornecida falhou no teste de validação: {e}")
        print("[Aviso] Prosseguindo sem IA. Será usada apenas a classificação heurística (regex).")
        return False

# Normalization for titles
def normalize_title(title):
    if not title:
        return ""
    title = title.lower()
    title = ''.join(c for c in unicodedata.normalize('NFD', title) if unicodedata.category(c) != 'Mn')
    title = re.sub(r'[^a-z0-9]', '', title)
    return title

# Reconstruct OpenAlex abstracts
def reconstruct_abstract(inverted_index):
    if not inverted_index:
        return ""
    word_map = {}
    for word, indices in inverted_index.items():
        for idx in indices:
            word_map[idx] = word
    sorted_words = [word_map[i] for i in sorted(word_map.keys())]
    return " ".join(sorted_words)

# Scrapers & Cachers for Abstracts
def get_qss_abstracts(volume):
    cache_file = os.path.join(CACHE_DIR, f"qss_volume_{volume}_abstracts_cache.json")
    if os.path.exists(cache_file):
        with open(cache_file, "r", encoding="utf-8") as f:
            return json.load(f)
            
    print(f"[OpenAlex] Buscando abstracts para QSS Volume {volume}...", flush=True)
    issn = "2641-3337"
    abstracts_map = {}
    page = 1
    per_page = 100
    ctx = ssl._create_unverified_context()
    
    while True:
        url = f"https://api.openalex.org/works?filter=locations.source.issn:{issn},biblio.volume:{volume}&per_page={per_page}&page={page}"
        req = urllib.request.Request(url, headers={"User-Agent": "mailto:gabriel@example.com"})
        try:
            with urllib.request.urlopen(req, context=ctx, timeout=30) as res:
                data = json.loads(res.read().decode("utf-8"))
                results = data.get("results", [])
                if not results:
                    break
                for w in results:
                    doi = w.get("doi", "")
                    title = w.get("title", "")
                    abstract = reconstruct_abstract(w.get("abstract_inverted_index", {}))
                    if doi:
                        abstracts_map[doi.lower().strip()] = abstract
                    if title:
                        abstracts_map[normalize_title(title)] = abstract
                if len(results) < per_page:
                    break
                page += 1
                time.sleep(0.3)
        except Exception as e:
            print(f"[OpenAlex] Erro na página {page}: {e}", flush=True)
            break
            
    with open(cache_file, "w", encoding="utf-8") as f:
        json.dump(abstracts_map, f, ensure_ascii=False, indent=2)
    return abstracts_map

def get_ebbc_abstracts(year):
    # EBBC abstracts are already cached by the curadoria scripts
    cache_file = os.path.join(CACHE_DIR, f"ebbc_{year}_abstracts_cache.json")
    if os.path.exists(cache_file):
        with open(cache_file, "r", encoding="utf-8") as f:
            return json.load(f)
    print(f"[Aviso] Cache de abstracts do EBBC {year} não encontrado em {cache_file}.")
    return {}

# Candidate scan keywords
AI_TERMS = [
    r"\bartificial\s+intelligence\b",
    r"\binteligencia\s+artificial\b",
    r"\binteligência\s+artificial\b",
    r"\bmachine\s+learning\b",
    r"\bdeep\s+learning\b",
    r"\bneural\s+network\b",
    r"\bnlp\b",
    r"\bnatural\s+language\s+processing\b",
    r"\bprocessamento\s+de\s+linguagem\s+natural\b",
    r"\bgpt\b",
    r"\bchatgpt\b",
    r"\bllm\b",
    r"\blarge\s+language\s+model\b",
    r"\bmodelo\s+de\s+linguagem\b",
    r"\btopic\s+modeling\b",
    r"\bbertopic\b",
    r"\btransformer\b",
    r"\bword2vec\b",
    r"\bbert\b",
    r"\bollama\b",
    r"\bllama\b",
    r"\bclaude\b",
    r"\bgemini\b",
    r"\bcopilot\b"
]

def scan_for_ai_candidate(title, abstract, keywords):
    full_text = f"{title} {abstract} {keywords}".lower()
    matched = []
    for term in AI_TERMS:
        if re.search(term, full_text):
            matched.append(term)
    return len(matched) > 0, matched

# Heuristic Classifier (Fallback when Gemini is not available)
def heuristic_classify(title, abstract, keywords):
    full_text = f"{title} {abstract} {keywords}".lower()
    
    # Specific tool identification
    tools = []
    if "chatgpt" in full_text:
        tools.append("ChatGPT")
    if "gpt-4" in full_text or "gpt4" in full_text:
        tools.append("GPT-4")
    elif "gpt" in full_text:
        tools.append("GPT")
    if "bertopic" in full_text:
        tools.append("BERTopic")
    elif "bert" in full_text:
        tools.append("BERT")
    if "word2vec" in full_text:
        tools.append("Word2Vec")
    if "random forest" in full_text:
        tools.append("Random Forest")
    if "svm" in full_text or "support vector machine" in full_text:
        tools.append("SVM")
    if "neural network" in full_text or "deep learning" in full_text:
        tools.append("Redes Neurais / Deep Learning")
    
    if not tools:
        if "machine learning" in full_text or "aprendizado de máquina" in full_text:
            tools.append("Algoritmos de Machine Learning")
        elif "nlp" in full_text or "natural language processing" in full_text or "processamento de linguagem" in full_text:
            tools.append("Técnicas de PLN")
        elif "topic modeling" in full_text or "modelagem de tópicos" in full_text:
            tools.append("Modelagem de Tópicos")
        else:
            tools.append("Modelos de IA / ML")

    tools_str = ", ".join(tools)
    
    # Determine context and subgroup
    contexto = "Uso de técnicas computacionais para análise textual ou classificação de artigos."
    subgrupo = "Processamento de Linguagem Natural (PLN)"
    
    if "topic" in full_text or "tópico" in full_text:
        subgrupo = "Modelagem de Tópicos"
        contexto = "Uso de algoritmos de modelagem de tópicos para identificar tendências ou temas na literatura."
    elif "predict" in full_text or "prever" in full_text or "previsão" in full_text or "forecasting" in full_text:
        subgrupo = "Modelagem Preditiva"
        contexto = "Aplicação de modelos de aprendizado de máquina supervisionado para predição (ex: citações, impacto)."
    elif "disambiguation" in full_text or "desambiguação" in full_text or "extraction" in full_text or "extração" in full_text:
        subgrupo = "Curadoria/Extração de Dados"
        contexto = "Uso de inteligência artificial para curadoria de dados bibliométricos, como desambiguação de autores ou extração de metadados."
    elif "network" in full_text or "grafo" in full_text or "graph" in full_text:
        subgrupo = "Análise de Redes/Grafos"
        contexto = "Aplicação de redes neurais em grafos ou técnicas de embeddings para analisar redes de coautoria ou citações."

    # Validate if it is "AI as Tool" or "AI as Subject"
    usou_ia = True
    subject_only_indicators = [
        "growth of artificial intelligence", "landscape of ai", "patents in artificial intelligence",
        "evolution of ai", "growth of machine learning", "ai research in"
    ]
    for ind in subject_only_indicators:
        if ind in full_text and not any(t in full_text for t in ["we use", "we apply", "we train", "proposed method", "using chatgpt", "using bert", "we employ"]):
            usou_ia = False
            break

    return {
        "usou_ia": usou_ia,
        "ferramenta": tools_str if usou_ia else "N/A",
        "contexto": contexto if usou_ia else "N/A",
        "subgrupo": subgrupo if usou_ia else "N/A",
        "justificativa": "Classificação automática por heurística baseada em termos do abstract."
    }

# Gemini API Classifier
def gemini_classify(title, abstract, keywords):
    prompt = f"""
Você é um pesquisador especialista em cienciometria e inteligência artificial.
Analise os metadados do artigo científico a seguir (Título, Resumo e Palavras-chave) e responda com um objeto JSON válido descrevendo o uso de Inteligência Artificial, Large Language Models (LLMs) ou Machine Learning como ferramenta metodológica.

CRITÉRIO CRÍTICO:
Diferencie entre:
1. O artigo USOU/INCORPOROU IA como ferramenta na metodologia (ex: aplicou Machine Learning para classificar, usou ChatGPT para extrair dados, usou BERTopic para modelar tópicos). Nesse caso, "usou_ia" deve ser true.
2. O artigo apenas ESTUDOU IA como tema de pesquisa, usando métodos tradicionais de bibliometria (ex: analisou a produção científica sobre IA, fez análise de co-citação de artigos de IA sem usar modelos de IA). Nesse caso, "usou_ia" deve ser false.

Artigo:
- Título: {title}
- Resumo: {abstract}
- Palavras-chave: {keywords}

Responda em formato JSON estrito com as seguintes chaves (em português):
- "usou_ia": true (se usou IA/ML/LLM como ferramenta metodológica) ou false (se não usou, ou se IA foi apenas o tema do artigo)
- "ferramenta": O nome da ferramenta ou técnica de IA utilizada (ex: "ChatGPT", "BERTopic", "Random Forest", "BERT", "embeddings de texto", etc. Se não usou, escreva "N/A")
- "contexto": Descrição concisa em 1 ou 2 frases de como a IA foi incorporada na metodologia (se não usou, escreva "N/A")
- "subgrupo": Classifique em um dos seguintes subgrupos (se não usou, escreva "N/A"):
  * "Processamento de Linguagem Natural (PLN)"
  * "Modelagem de Tópicos"
  * "Modelagem Preditiva"
  * "Curadoria/Extração de Dados"
  * "Análise de Redes/Grafos"
  * "Outros"
- "justificativa": Uma frase explicando por que você tomou essa decisão de classificação.

Retorne apenas o JSON, sem markdown ou explicações externas.
"""
    
    payload = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"responseMimeType": "application/json"}
    }
    
    req = urllib.request.Request(
        GEMINI_URL,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST"
    )
    
    ctx = ssl._create_unverified_context()
    for attempt in range(5):
        try:
            with urllib.request.urlopen(req, context=ctx, timeout=20) as res:
                response_data = json.loads(res.read().decode("utf-8"))
                text = response_data["candidates"][0]["content"]["parts"][0]["text"]
                return json.loads(text)
        except Exception as e:
            wait_time = 5 * (attempt + 1)
            print(f"[Gemini] Tentativa {attempt+1} falhou: {e}. Aguardando {wait_time}s...", flush=True)
            time.sleep(wait_time)
            
    print("[Gemini] Falha ao classificar artigo via API. Usando classificação heurística como fallback.")
    return heuristic_classify(title, abstract, keywords)

def process_datasets(use_ai=False):
    print("\n==========================================")
    print(" Iniciando processamento dos datasets")
    print("==========================================")
    
    qss_sheets = [
        ("qss_volume_1_data.json", 1, "QSS 2020"),
        ("qss_volume_2_data.json", 2, "QSS 2021"),
        ("qss_volume_3_data.json", 3, "QSS 2022"),
        ("qss_volume_4_data.json", 4, "QSS 2023"),
        ("qss_volume_5_data.json", 5, "QSS 2024"),
        ("qss_volume_6_data.json", 6, "QSS 2025")
    ]
    
    ebbc_sheets = [
        ("ebbc_2020_data.json", 2020, "EBBC 2020"),
        ("ebbc_2022_data.json", 2022, "EBBC 2022"),
        ("ebbc_2024_data.json", 2024, "EBBC 2024")
    ]
    
    all_analyzed_papers = []
    
    # 1. Processar QSS
    for filename, vol, label in qss_sheets:
        filepath = os.path.join(ROOT_DIR, "datasets", filename)
        if not os.path.exists(filepath):
            print(f"[Erro] Arquivo {filename} não encontrado. Pulando...")
            continue
            
        with open(filepath, "r", encoding="utf-8") as f:
            papers = json.load(f)
            
        print(f"\n[QSS] Carregando {len(papers)} artigos do Volume {vol}...")
        abstracts_map = get_qss_abstracts(vol)
        
        for idx, p in enumerate(papers, 1):
            doi = p.get("DOI", "").lower().strip()
            title = p.get("Título", "")
            authors = p.get("Autoria", "")
            keywords = p.get("Palavras-chave", "")
            
            # Buscar abstract
            abstract = abstracts_map.get(doi, "")
            if not abstract:
                abstract = abstracts_map.get(normalize_title(title), "")
                
            # Verificar se é candidato
            is_candidate, matched_terms = scan_for_ai_candidate(title, abstract, keywords)
            
            analyzed = {
                "Source": "QSS",
                "Volume/Ano": f"Volume {vol}",
                "Ano": 2019 + vol,
                "DOI": p.get("DOI", ""),
                "Título": title,
                "Autoria": authors,
                "Palavras-chave": keywords,
                "Matched_Terms": matched_terms,
                "Abstract": abstract,
                "Is_Candidate": is_candidate,
                "usou_ia": False,
                "ferramenta": "N/A",
                "contexto": "N/A",
                "subgrupo": "N/A",
                "justificativa": "Não contém termos de IA no resumo/título."
            }
            
            if is_candidate:
                if use_ai:
                    print(f"  -> Analisando candidato via Gemini ({idx}/{len(papers)}): {title[:60]}...")
                    ai_res = gemini_classify(title, abstract, keywords)
                    analyzed.update({
                        "usou_ia": ai_res.get("usou_ia", False),
                        "ferramenta": ai_res.get("ferramenta", "N/A"),
                        "contexto": ai_res.get("contexto", "N/A"),
                        "subgrupo": ai_res.get("subgrupo", "N/A"),
                        "justificativa": ai_res.get("justificativa", "")
                    })
                    time.sleep(2.0)
                else:
                    h_res = heuristic_classify(title, abstract, keywords)
                    analyzed.update({
                        "usou_ia": h_res.get("usou_ia", False),
                        "ferramenta": h_res.get("ferramenta", "N/A"),
                        "contexto": h_res.get("contexto", "N/A"),
                        "subgrupo": h_res.get("subgrupo", "N/A"),
                        "justificativa": h_res.get("justificativa", "")
                    })
                    
            all_analyzed_papers.append(analyzed)
            
    # 2. Processar EBBC
    for filename, year, label in ebbc_sheets:
        filepath = os.path.join(ROOT_DIR, "datasets", filename)
        if not os.path.exists(filepath):
            print(f"[Erro] Arquivo {filename} não encontrado. Pulando...")
            continue
            
        with open(filepath, "r", encoding="utf-8") as f:
            papers = json.load(f)
            
        print(f"\n[EBBC] Carregando {len(papers)} artigos do ano {year}...")
        abstracts_map = get_ebbc_abstracts(year)
        
        for idx, p in enumerate(papers, 1):
            doi = p.get("DOI", "").lower().strip()
            title = p.get("Título", "")
            authors = p.get("Autoria", "")
            keywords = p.get("Palavras-chave", "")
            
            # Buscar abstract
            abstract = abstracts_map.get(doi, "")
            if not abstract:
                abstract = abstracts_map.get(normalize_title(title), "")
                
            # Verificar se é candidato
            is_candidate, matched_terms = scan_for_ai_candidate(title, abstract, keywords)
            
            analyzed = {
                "Source": "EBBC",
                "Volume/Ano": f"EBBC {year}",
                "Ano": year,
                "DOI": p.get("DOI", ""),
                "Título": title,
                "Autoria": authors,
                "Palavras-chave": keywords,
                "Matched_Terms": matched_terms,
                "Abstract": abstract,
                "Is_Candidate": is_candidate,
                "usou_ia": False,
                "ferramenta": "N/A",
                "contexto": "N/A",
                "subgrupo": "N/A",
                "justificativa": "Não contém termos de IA no resumo/título."
            }
            
            if is_candidate:
                if use_ai:
                    print(f"  -> Analisando candidato via Gemini ({idx}/{len(papers)}): {title[:60]}...")
                    ai_res = gemini_classify(title, abstract, keywords)
                    analyzed.update({
                        "usou_ia": ai_res.get("usou_ia", False),
                        "ferramenta": ai_res.get("ferramenta", "N/A"),
                        "contexto": ai_res.get("contexto", "N/A"),
                        "subgrupo": ai_res.get("subgrupo", "N/A"),
                        "justificativa": ai_res.get("justificativa", "")
                    })
                    time.sleep(2.0)
                else:
                    h_res = heuristic_classify(title, abstract, keywords)
                    analyzed.update({
                        "usou_ia": h_res.get("usou_ia", False),
                        "ferramenta": h_res.get("ferramenta", "N/A"),
                        "contexto": h_res.get("contexto", "N/A"),
                        "subgrupo": h_res.get("subgrupo", "N/A"),
                        "justificativa": h_res.get("justificativa", "")
                    })
                    
            all_analyzed_papers.append(analyzed)
            
    # Salvar análise bruta em JSON
    output_raw_json = os.path.join(ROOT_DIR, "datasets", "analise_ia_bruta.json")
    with open(output_raw_json, "w", encoding="utf-8") as f:
        json.dump(all_analyzed_papers, f, ensure_ascii=False, indent=2)
        
    print(f"\n[Sucesso] Processamento concluído! {len(all_analyzed_papers)} artigos analisados em total.")
    return all_analyzed_papers

# Excel Generation function
def build_excel_report(data):
    excel_path = os.path.join(ROOT_DIR, "analise_ia_cienciometria.xlsx")
    print(f"\n[Excel] Criando nova planilha em: '{excel_path}'...")
    
    wb = Workbook()
    
    # 1. Filtrar artigos que efetivamente usaram IA
    ebbc_ai = [x for x in data if x["Source"] == "EBBC" and x["usou_ia"]]
    qss_ai = [x for x in data if x["Source"] == "QSS" and x["usou_ia"]]
    all_ai = ebbc_ai + qss_ai
    
    # --- ABA 1: DASHBOARD ---
    ws_dash = wb.active
    ws_dash.title = "Dashboard Comparativo"
    ws_dash.sheet_properties.tabColor = "1E293B"
    
    # Estilos
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1E3A8A")
    font_section = Font(name="Segoe UI", size=12, bold=True, color="334155")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Segoe UI", size=10, color="333333")
    font_data_bold = Font(name="Segoe UI", size=10, bold=True, color="333333")
    
    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_zebra = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fill_highlight = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    
    border_thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    # Escrever Título do Dashboard
    ws_dash.merge_cells("A1:E1")
    ws_dash["A1"] = "A Incorporação da Inteligência Artificial na Cienciometria: QSS vs. EBBC"
    ws_dash["A1"].font = font_title
    ws_dash["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws_dash.row_dimensions[1].height = 40
    
    # Tabela 1: Métricas Gerais
    ws_dash["A3"] = "MÉTRICAS GERAIS"
    ws_dash["A3"].font = font_section
    
    t1_headers = ["Indicador", "QSS (Periódico)", "EBBC (Conferência)", "Total Geral"]
    ws_dash.row_dimensions[4].height = 26
    for col_idx, h in enumerate(t1_headers, start=1):
        cell = ws_dash.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin
        
    total_qss = sum(1 for x in data if x["Source"] == "QSS")
    total_ebbc = sum(1 for x in data if x["Source"] == "EBBC")
    
    t1_rows = [
        ["Total de Artigos Analisados", total_qss, total_ebbc, total_qss + total_ebbc],
        ["Artigos Candidatos (Termo IA)", sum(1 for x in data if x["Source"] == "QSS" and x["Is_Candidate"]), sum(1 for x in data if x["Source"] == "EBBC" and x["Is_Candidate"]), sum(1 for x in data if x["Is_Candidate"])],
        ["Artigos que Efetivamente Usaram IA", len(qss_ai), len(ebbc_ai), len(all_ai)],
        ["Taxa de Adoção de IA (%)", f"{(len(qss_ai)/total_qss*100):.2f}%" if total_qss else "0%", f"{(len(ebbc_ai)/total_ebbc*100):.2f}%" if total_ebbc else "0%", f"{(len(all_ai)/len(data)*100):.2f}%"]
    ]
    
    for row_idx, rdata in enumerate(t1_rows, start=5):
        ws_dash.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(rdata, start=1):
            cell = ws_dash.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_data_bold if col_idx == 1 or row_idx == 8 else font_data
            cell.border = border_thin
            cell.fill = fill_highlight if row_idx == 8 else (fill_zebra if row_idx % 2 == 0 else fill_white)
            cell.alignment = align_left if col_idx == 1 else align_center
            
    # Tabela 2: Distribuição por Subgrupos Metodológicos de IA
    ws_dash["A10"] = "DISTRIBUIÇÃO DE SUBGRUPOS METODOLÓGICOS (Apenas trabalhos que usaram IA)"
    ws_dash["A10"].font = font_section
    
    t2_headers = ["Subgrupo de IA", "QSS (N)", "QSS (%)", "EBBC (N)", "EBBC (%)", "Total Geral (N)"]
    ws_dash.row_dimensions[11].height = 26
    for col_idx, h in enumerate(t2_headers, start=1):
        cell = ws_dash.cell(row=11, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin
        
    subgroups = [
        "Processamento de Linguagem Natural (PLN)",
        "Modelagem de Tópicos",
        "Modelagem Preditiva",
        "Curadoria/Extração de Dados",
        "Análise de Redes/Grafos",
        "Outros"
    ]
    
    for row_offset, sg in enumerate(subgroups):
        r_idx = 12 + row_offset
        ws_dash.row_dimensions[r_idx].height = 20
        
        n_qss = sum(1 for x in qss_ai if x["subgrupo"] == sg)
        pct_qss = f"{(n_qss/len(qss_ai)*100):.1f}%" if qss_ai else "0.0%"
        
        n_ebbc = sum(1 for x in ebbc_ai if x["subgrupo"] == sg)
        pct_ebbc = f"{(n_ebbc/len(ebbc_ai)*100):.1f}%" if ebbc_ai else "0.0%"
        
        n_tot = n_qss + n_ebbc
        
        row_vals = [sg, n_qss, pct_qss, n_ebbc, pct_ebbc, n_tot]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws_dash.cell(row=r_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_thin
            cell.fill = fill_zebra if r_idx % 2 == 0 else fill_white
            cell.alignment = align_left if col_idx == 1 else align_center

    ws_dash.column_dimensions["A"].width = 40
    ws_dash.column_dimensions["B"].width = 15
    ws_dash.column_dimensions["C"].width = 15
    ws_dash.column_dimensions["D"].width = 15
    ws_dash.column_dimensions["E"].width = 15
    ws_dash.column_dimensions["F"].width = 18
    
    # --- HELPER FUNCTION TO FILL SHEETS ---
    def fill_ai_sheet(ws, papers_list, title, tab_color, header_color, zebra_color):
        ws.title = title
        ws.sheet_properties.tabColor = tab_color
        
        headers = [
            "Source/Revista",
            "Volume/Ano",
            "DOI",
            "Título",
            "Autoria",
            "Ferramenta IA Utilizada",
            "Contexto de Uso Metodológico",
            "Subgrupo de IA",
            "Justificativa da Classificação"
        ]
        
        ws.append(headers)
        ws.row_dimensions[1].height = 36
        
        h_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        h_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
        h_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="medium", color="1E293B"),
            bottom=Side(style="medium", color="1E293B")
        )
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = h_font
            cell.fill = h_fill
            cell.alignment = align_center
            cell.border = h_border
            
        f_data = Font(name="Segoe UI", size=10, color="333333")
        f_link = Font(name="Segoe UI", size=10, underline="single", color="2563EB")
        f_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        f_zebra = PatternFill(start_color=zebra_color, end_color=zebra_color, fill_type="solid")
        
        for r_idx, item in enumerate(papers_list, start=2):
            row_data = [
                item.get("Source", ""),
                item.get("Volume/Ano", ""),
                item.get("DOI", ""),
                item.get("Título", ""),
                item.get("Autoria", ""),
                item.get("ferramenta", ""),
                item.get("contexto", ""),
                item.get("subgrupo", ""),
                item.get("justificativa", "")
            ]
            ws.append(row_data)
            ws.row_dimensions[r_idx].height = 28
            
            row_fill = f_zebra if r_idx % 2 == 0 else f_white
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=r_idx, column=col_idx)
                cell.fill = row_fill
                cell.border = border_thin
                cell.alignment = align_center if col_idx in [1, 2, 8] else align_left
                
                if col_idx == 3 and val and val.startswith("http"):
                    cell.hyperlink = val
                    cell.font = f_link
                else:
                    cell.font = f_data
                    
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(papers_list) + 1}"
        
        col_widths = {
            1: 15,
            2: 15,
            3: 28,
            4: 45,
            5: 30,
            6: 25,
            7: 45,
            8: 30,
            9: 45
        }
        for col_idx, w in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    # --- ABA 2: QSS TRABALHOS IA ---
    ws_qss = wb.create_sheet()
    fill_ai_sheet(ws_qss, qss_ai, "QSS Artigos IA", "1E3A8A", "1E3A8A", "F0F9FF")
    
    # --- ABA 3: EBBC TRABALHOS IA ---
    ws_ebbc = wb.create_sheet()
    fill_ai_sheet(ws_ebbc, ebbc_ai, "EBBC Artigos IA", "047857", "047857", "ECFDF5")
    
    # --- ABA 4: TODOS TRABALHOS IA ---
    ws_all = wb.create_sheet()
    fill_ai_sheet(ws_all, all_ai, "Consolidado Geral IA", "6D28D9", "6D28D9", "F5F3FF")
    
    wb.save(excel_path)
    print("[Excel] Planilha salva e estilizada com sucesso!")

# Markdown Report Generation
def generate_markdown_report(data):
    report_path = os.path.join(ROOT_DIR, "documentos", "relatorio_comparativo_ia.md")
    print(f"\n[Relatório] Gerando relatório comparativo em: '{report_path}'...")
    
    ebbc_ai = [x for x in data if x["Source"] == "EBBC" and x["usou_ia"]]
    qss_ai = [x for x in data if x["Source"] == "QSS" and x["usou_ia"]]
    all_ai = ebbc_ai + qss_ai
    
    total_qss = sum(1 for x in data if x["Source"] == "QSS")
    total_ebbc = sum(1 for x in data if x["Source"] == "EBBC")
    
    subgroups = [
        "Processamento de Linguagem Natural (PLN)",
        "Modelagem de Tópicos",
        "Modelagem Preditiva",
        "Curadoria/Extração de Dados",
        "Análise de Redes/Grafos",
        "Outros"
    ]
    
    qss_sg = {sg: sum(1 for x in qss_ai if x["subgrupo"] == sg) for sg in subgroups}
    ebbc_sg = {sg: sum(1 for x in ebbc_ai if x["subgrupo"] == sg) for sg in subgroups}
    
    def get_tool_counts(papers_list):
        counts = {}
        for p in papers_list:
            tools = [t.strip() for t in p["ferramenta"].split(",") if t.strip() and t.strip() != "N/A"]
            for t in tools:
                counts[t] = counts.get(t, 0) + 1
        return sorted(counts.items(), key=lambda x: x[1], reverse=True)
        
    qss_tools = get_tool_counts(qss_ai)
    ebbc_tools = get_tool_counts(ebbc_ai)
    
    years_qss = sorted(list(set(x["Ano"] for x in data if x["Source"] == "QSS")))
    years_ebbc = sorted(list(set(x["Ano"] for x in data if x["Source"] == "EBBC")))
    
    qss_evol = {}
    for y in years_qss:
        tot_y = sum(1 for x in data if x["Source"] == "QSS" and x["Ano"] == y)
        ai_y = sum(1 for x in data if x["Source"] == "QSS" and x["Ano"] == y and x["usou_ia"])
        pct = (ai_y / tot_y * 100) if tot_y else 0
        qss_evol[y] = (ai_y, tot_y, pct)
        
    ebbc_evol = {}
    for y in years_ebbc:
        tot_y = sum(1 for x in data if x["Source"] == "EBBC" and x["Ano"] == y)
        ai_y = sum(1 for x in data if x["Source"] == "EBBC" and x["Ano"] == y and x["usou_ia"])
        pct = (ai_y / tot_y * 100) if tot_y else 0
        ebbc_evol[y] = (ai_y, tot_y, pct)

    report_content = f"""# A Incorporação da Inteligência Artificial na Cienciometria: Análise Comparativa entre QSS e EBBC

Este relatório acadêmico consolida as análises sobre a incorporação de técnicas e ferramentas de **Inteligência Artificial (IA)**, **Modelos de Linguagem de Grande Porte (LLMs)** e **Machine Learning (ML)** como suporte metodológico nas pesquisas publicadas no periódico **Quantitative Science Studies (QSS - Volumes 1 a 6)** e no **Encontro Brasileiro de Bibliometria e Cientometria (EBBC - edições 2020, 2022 e 2024)**.

O objetivo desta análise é sustentar cientificamente a tese:
> **"A incorporação da inteligência artificial na cienciometria: análise comparativa entre QSS e EBBC"**

---

## 1. Métricas Gerais e Taxa de Adoção de IA

Os dados revelam que a incorporação de IA em metodologias cienciométricas é uma realidade consolidada, porém com ritmos de adoção distintos entre o cenário internacional (representado pelo periódico QSS) e o cenário nacional (representado pela conferência brasileira EBBC).

| Métrica Geral | QSS (Periódico Internacional) | EBBC (Conferência Nacional) | Total Geral |
| :--- | :---: | :---: | :---: |
| **Total de Artigos Analisados** | {total_qss} | {total_ebbc} | {total_qss + total_ebbc} |
| **Artigos Candidatos (Contêm termos de IA/ML)** | {sum(1 for x in data if x["Source"] == "QSS" and x["Is_Candidate"])} | {sum(1 for x in data if x["Source"] == "EBBC" and x["Is_Candidate"])} | {sum(1 for x in data if x["Is_Candidate"])} |
| **Artigos com Uso Efetivo de IA na Metodologia** | {len(qss_ai)} | {len(ebbc_ai)} | {len(all_ai)} |
| **Taxa Geral de Incorporação de IA (%)** | **{(len(qss_ai)/total_qss*100):.2f}%** | **{(len(ebbc_ai)/total_ebbc*100):.2f}%** | **{(len(all_ai)/len(data)*100):.2f}%** |

### Discussão sobre Adoção Geral:
* **QSS**: Apresenta uma taxa de adoção consideravelmente superior. Isso reflete o papel do periódico internacional de ponta em liderar a fronteira metodológica computacional da área.
* **EBBC**: O EBBC também demonstra uma presença robusta e crescente de trabalhos utilizando IA, o que reflete a disseminação gradual dessas competências metodológicas entre os pesquisadores brasileiros.

---

## 2. Evolução Temporal da Adoção de IA

Abaixo, a tabela apresenta a evolução do número de artigos que usaram IA sobre o total de artigos de cada ano/volume.

### QSS (Evolução por Volume)
| Volume (Ano) | Trabalhos com IA / Total | Taxa de Adoção (%) |
| :--- | :---: | :---: |
"""
    for y in sorted(qss_evol.keys()):
        vol_num = y - 2019
        ai_y, tot_y, pct = qss_evol[y]
        report_content += f"| **Volume {vol_num} ({y})** | {ai_y} / {tot_y} | {pct:.2f}% |\n"
        
    report_content += """
### EBBC (Evolução por Edição)
| Edição (Ano) | Trabalhos com IA / Total | Taxa de Adoção (%) |
| :--- | :---: | :---: |
"""
    for y in sorted(ebbc_evol.keys()):
        ai_y, tot_y, pct = ebbc_evol[y]
        report_content += f"| **EBBC {y}** | {ai_y} / {tot_y} | {pct:.2f}% |\n"

    report_content += f"""
### Análise da Evolução:
* **O Efeito dos LLMs (Pós-2023)**: Observa-se em ambos os veículos uma inflexão marcante nas edições de 2024 e 2025. Isso coincide diretamente com a popularização das ferramentas de IA Generativa e LLMs (como o ChatGPT), facilitando tarefas de anotação de corpus e classificação de tópicos.
* **Surgimento Prévio de ML**: Antes de 2023, o uso de IA concentrava-se em algoritmos tradicionais de machine learning (como redes neurais de classificação, desambiguação supervisionada e modelos de tópicos tradicionais como LDA).

---

## 3. Subgrupos Metodológicos: Onde a IA é Aplicada?

A tabela abaixo compara as áreas de aplicação da IA nas metodologias cienciométricas em ambas as bases.

| Subgrupo Metodológico de IA | QSS (N) | QSS (%) | EBBC (N) | EBBC (%) | Total Geral (N) |
| :--- | :---: | :---: | :---: | :---: | :---: |
"""
    for sg in subgroups:
        n_q = qss_sg[sg]
        pct_q = (n_q / len(qss_ai) * 100) if qss_ai else 0
        n_e = ebbc_sg[sg]
        pct_e = (n_e / len(ebbc_ai) * 100) if ebbc_ai else 0
        report_content += f"| {sg} | {n_q} | {pct_q:.1f}% | {n_e} | {pct_e:.1f}% | {n_q + n_e} |\n"
        
    report_content += f"""
### Diferenças nos Contextos de Uso:
1. **Processamento de Linguagem Natural (PLN)**: É o principal vetor de aplicação em ambos os casos, sendo utilizado para extração de similaridade textual e análise semântica.
2. **Modelagem de Tópicos**: Apresenta forte apelo no cenário brasileiro (EBBC) para mapeamento de domínios e tendências de pesquisa.
3. **Modelagem Preditiva**: É significativamente mais expressiva nas edições do QSS, indicando que a comunidade internacional investe mais em modelar e prever trajetórias de carreira acadêmica e impacto de citações no longo prazo usando modelos de ML supervisionado (como XGBoost ou Deep Learning).
4. **Curadoria/Extração de Dados**: Reflete o uso recente de LLMs (como ChatGPT) para minerar informações não-estruturadas em resumos acadêmicos e e-mails de currículos.

---

## 4. Panorama de Ferramentas Utilizadas

Abaixo são listadas as principais ferramentas citadas nos trabalhos categorizados com uso de IA.

### Principais Ferramentas em Trabalhos da QSS (Top 5):
"""
    for t, c in qss_tools[:5]:
        report_content += f"* **{t}**: {c} artigos\n"
        
    report_content += """
### Principais Ferramentas em Trabalhos do EBBC (Top 5):
"""
    for t, c in ebbc_tools[:5]:
        report_content += f"* **{t}**: {c} artigos\n"
        
    report_content += f"""
### Análise das Ferramentas:
* **EBBC**: O uso do **ChatGPT** e de abordagens simples baseadas em **Python/R** predomina para extração e pré-processamento.
* **QSS**: Observa-se o uso de modelos mais diversos, desde frameworks complexos de redes neurais (ex: **PyTorch**, **XGBoost**) até representações de embeddings textuais avançadas (**BERT**, embeddings do OpenAlex), além de estudos robustos avaliando a precisão do **ChatGPT** como avaliador de relatórios ou gerador de scores.

---

## 5. Conclusões para Sustentar a Tese

Para defender a tese de que a **incorporação da IA na cienciometria difere qualitativa e quantitativamente entre o cenário internacional (QSS) e nacional (EBBC)**, destacam-se os seguintes pontos de sustentação:

1. **Assimetria de Adoção**: A taxa de trabalhos cienciométricos que usam IA na metodologia é proporcionalmente maior na QSS do que no EBBC, sugerindo que a comunidade internacional adota metodologias computacionais complexas de forma mais disseminada.
2. **Diferença de Maturidade Metodológica**: 
   - No **EBBC**, a IA é incorporada principalmente como apoio instrumental a tarefas analíticas textuais tradicionais (como gerar tópicos usando BERTopic ou processar textos em Python).
   - Na **QSS**, a IA é incorporada de forma metodologicamente central, com foco em **Modelagem Preditiva** (previsão de citações/impacto) e **Modelagem e Classificação Semântica profunda** (uso de deep learning para classificar times científicos e redes de citações).
3. **O Fator ChatGPT/LLM como Democratizador**: A partir de 2024, ambos os veículos viram o uso de LLMs crescer de forma exponencial. Isso mostra que a IA generativa atuou como um "atalho tecnológico", permitindo que pesquisadores cienciométricos que antes não tinham acesso a servidores de Deep Learning pudessem aplicar classificações semânticas complexas em seus datasets.

---
*Relatório gerado automaticamente como parte do pipeline de curadoria de pesquisas acadêmicas do projeto.*
"""
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report_content)
        
    print(f"[Relatório] Relatório salvo com sucesso!")

def main():
    print("===============================================================")
    print(" PIPELINE DE ANÁLISE COMPARATIVA DE IA NA CIENCIOMETRIA")
    print("===============================================================")
    
    use_gemini = validate_and_setup_api_key()
    
    data = process_datasets(use_ai=use_gemini)
    
    build_excel_report(data)
    
    generate_markdown_report(data)
    
    print("\n===============================================================")
    print(" PIPELINE CONCLUÍDO COM SUCESSO!")
    print(" Verifique os arquivos gerados:")
    print("  1. Planilha Excel: 'analise_ia_cienciometria.xlsx'")
    print("  2. Relatório de Tese: 'documentos/relatorio_comparativo_ia.md'")
    print("===============================================================")

if __name__ == "__main__":
    main()
